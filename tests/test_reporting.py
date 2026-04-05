from __future__ import annotations

from dataclasses import replace
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from trading_lab.contracts import (  # noqa: E402
    BacktestResult,
    BacktestSpec,
    CostAssumptions,
    InstrumentMeta,
    NodeContract,
    NodeSpec,
)
from trading_lab.experiments import (  # noqa: E402
    ExperimentSpec,
    FoldSpec,
    HoldoutSpec,
    PruningConfig,
    SearchConfig,
    VariantSpec,
)
from trading_lab.reporting import (  # noqa: E402
    build_price_plot_frame,
    build_summary_frames,
    build_trade_log_frame,
    export_deep_dive_artifacts,
    export_summary_workbook,
    select_deep_dive_targets,
)
from trading_lab.runner import PruneRecord, run_experiment  # noqa: E402


class FakeClock:
    def __init__(self, runtimes: list[float]) -> None:
        self.current = 0.0
        self._runtimes = list(runtimes)
        self._pending_runtime: float | None = None

    def __call__(self) -> float:
        if self._pending_runtime is None:
            self._pending_runtime = self._runtimes.pop(0) if self._runtimes else 0.0
            return self.current
        self.current += self._pending_runtime
        self._pending_runtime = None
        return self.current


def _instrument() -> InstrumentMeta:
    return InstrumentMeta(
        symbol="TEST",
        price_increment=0.01,
        quantity_increment=1.0,
    )


def _node_contract(
    name: str,
    kind: str,
    *,
    parameters: dict[str, object],
) -> NodeContract:
    emitted_actions = {
        "entry": ("enter_long", "hold"),
        "exit": ("close", "hold"),
        "risk": ("hold",),
    }[kind]
    return NodeContract(
        spec=NodeSpec(
            name=name,
            kind=kind,  # type: ignore[arg-type]
            emitted_action_types=emitted_actions,  # type: ignore[arg-type]
            required_history=2,
        ),
        manifest={"module": "tests.test_reporting", "parameters": parameters},
    )


def _variant(name: str, *, fast: int) -> VariantSpec:
    entry_name = f"{name}_entry"
    exit_name = f"{name}_exit"
    risk_name = f"{name}_risk"
    return VariantSpec(
        backtest_spec=BacktestSpec(
            name=name,
            instrument=_instrument(),
            entry_node=entry_name,
            exit_node=exit_name,
            risk_node=risk_name,
            initial_cash=100_000.0,
            costs=CostAssumptions(fee_rate=0.001),
        ),
        entry_contract=_node_contract(
            entry_name,
            "entry",
            parameters={"family": "sma_cross", "fast": fast, "slow": 20},
        ),
        exit_contract=_node_contract(
            exit_name,
            "exit",
            parameters={"family": "time_stop", "bars": 5},
        ),
        risk_contract=_node_contract(
            risk_name,
            "risk",
            parameters={"family": "fixed_fraction", "risk_fraction": 0.02},
        ),
    )


def _data() -> pd.DataFrame:
    rows = []
    for day, price in enumerate(range(100, 106), start=1):
        ts = pd.Timestamp(datetime(2024, 1, day))
        rows.append(
            {
                "ts": ts,
                "symbol": "TEST",
                "timeframe": "1D",
                "open": float(price),
                "high": float(price + 1),
                "low": float(price - 1),
                "close": float(price + 0.5),
                "volume": 1_000.0,
            }
        )
    return pd.DataFrame(rows)


def _label_from_slice(data_slice: pd.DataFrame) -> str:
    ts = pd.Timestamp(data_slice["ts"].min())
    mapping = {
        pd.Timestamp("2024-01-02"): "fold_a",
        pd.Timestamp("2024-01-03"): "fold_b",
        pd.Timestamp("2024-01-04"): "holdout",
    }
    return mapping[ts]


def _stub_result(
    spec: BacktestSpec,
    *,
    label: str,
    run_id: str,
    metrics: dict[str, object],
    final_equity: float = 100_000.0,
) -> BacktestResult:
    ts_map = {
        "fold_a": pd.Timestamp("2024-01-02 00:00:00"),
        "fold_b": pd.Timestamp("2024-01-03 00:00:00"),
        "holdout": pd.Timestamp("2024-01-04 00:00:00"),
    }
    trade_side = {"fold_a": "long", "fold_b": "short", "holdout": "long"}[label]
    trade_entry_price = {"fold_a": 101.0, "fold_b": 102.0, "holdout": 103.0}[label]
    trade_exit_price = {"fold_a": 102.0, "fold_b": 101.0, "holdout": 104.0}[label]
    ts = ts_map[label]
    trade_net_pnl = float(metrics.get("net_pnl", 0.0))
    return BacktestResult(
        spec=spec,
        decision_log=pd.DataFrame(),
        order_log=pd.DataFrame(),
        fill_log=pd.DataFrame(),
        trade_ledger=pd.DataFrame(
            [
                {
                    "trade_id": f"{run_id}-trade-1",
                    "symbol": spec.instrument.symbol,
                    "side": trade_side,
                    "entry_ts": ts,
                    "exit_ts": ts,
                    "qty": 1.0,
                    "entry_price": trade_entry_price,
                    "exit_price": trade_exit_price,
                    "gross_pnl": trade_net_pnl,
                    "net_pnl": trade_net_pnl,
                    "bars_held": 1,
                    "exit_reason": "stub-exit",
                    "fees": 0.0,
                }
            ]
        ),
        equity_curve=pd.DataFrame(
            [
                {
                    "ts": ts,
                    "cash": final_equity,
                    "equity": final_equity,
                    "realized_pnl": trade_net_pnl,
                    "unrealized_pnl": 0.0,
                    "gross_exposure": 0.0,
                    "net_exposure": 0.0,
                    "drawdown": 0.0,
                }
            ]
        ),
        artifacts={"run_manifest": {"run_id": run_id}, "stub_metrics": metrics},
    )


def _metrics_fn(result: BacktestResult) -> dict[str, object]:
    metrics = dict(result.artifacts["stub_metrics"])  # type: ignore[index]
    metrics["holding_period_distribution"] = {1: 1}
    return metrics


def _experiment_result():
    variants = (_variant("variant_a", fast=5), _variant("variant_b", fast=10))
    experiment = ExperimentSpec(
        name="reporting-phase",
        variants=variants,
        folds=(
            FoldSpec(
                fold_index=0,
                train_start=datetime(2024, 1, 1),
                train_end=datetime(2024, 1, 2),
                validation_start=datetime(2024, 1, 2),
                validation_end=datetime(2024, 1, 3),
                label="fold_a",
            ),
            FoldSpec(
                fold_index=1,
                train_start=datetime(2024, 1, 2),
                train_end=datetime(2024, 1, 3),
                validation_start=datetime(2024, 1, 3),
                validation_end=datetime(2024, 1, 4),
                label="fold_b",
            ),
        ),
        holdout=HoldoutSpec(
            start=datetime(2024, 1, 4),
            end=datetime(2024, 1, 5),
            label="holdout",
        ),
        search=SearchConfig(mode="grid", max_variants=2),
        pruning=PruningConfig(min_trades=3),
    )
    metrics_map = {
        ("variant_a", "fold_a"): {"trade_count": 2, "net_pnl": 10.0, "gross_pnl": 10.0},
        ("variant_a", "fold_b"): {"trade_count": 2, "net_pnl": 12.0, "gross_pnl": 12.0},
        ("variant_a", "holdout"): {"trade_count": 1, "net_pnl": 3.0, "gross_pnl": 3.0},
        ("variant_b", "fold_a"): {"trade_count": 1, "net_pnl": 1.0, "gross_pnl": 1.0},
        ("variant_b", "fold_b"): {"trade_count": 1, "net_pnl": 1.0, "gross_pnl": 1.0},
    }

    def backtest_fn(
        spec: BacktestSpec,
        data_slice: pd.DataFrame,
        *,
        node_registry=None,
    ) -> BacktestResult:
        label = _label_from_slice(data_slice)
        return _stub_result(
            spec,
            label=label,
            run_id=f"{spec.name}-{label}",
            metrics=metrics_map[(spec.name, label)],
        )

    return run_experiment(
        experiment,
        _data(),
        backtest_fn=backtest_fn,
        metrics_fn=_metrics_fn,
        time_fn=FakeClock([1.0] * 5),
    )


def test_build_summary_frames_include_required_sheets_and_columns() -> None:
    result = _experiment_result()

    frames = build_summary_frames(result)

    assert tuple(frames) == (
        "run_summary",
        "variant_summary",
        "fold_metrics",
        "holdout_metrics",
        "failures_prunes",
        "config",
    )
    assert {
        "variant_id",
        "status",
        "cv_runtime_seconds",
        "entry_node",
        "entry_params",
        "risk_params",
    }.issubset(frames["variant_summary"].columns)
    assert {
        "variant_id",
        "fold_label",
        "variant_status",
        "fold_status",
        "runtime_seconds",
    }.issubset(frames["fold_metrics"].columns)
    assert {
        "variant_id",
        "holdout_label",
        "variant_status",
        "holdout_status",
        "runtime_seconds",
    }.issubset(frames["holdout_metrics"].columns)
    assert {
        "record_type",
        "variant_id",
        "reason",
        "message",
    }.issubset(frames["failures_prunes"].columns)


def test_export_summary_workbook_generates_required_sheets(tmp_path: Path) -> None:
    result = _experiment_result()
    output_path = tmp_path / "experiment_summary.xlsx"

    written_path = export_summary_workbook(result, output_path)

    assert written_path == output_path
    assert output_path.exists()

    workbook = load_workbook(output_path, read_only=True)
    assert workbook.sheetnames == [
        "run_summary",
        "variant_summary",
        "fold_metrics",
        "holdout_metrics",
        "failures_prunes",
        "config",
    ]


def test_summary_workbook_has_required_columns(tmp_path: Path) -> None:
    result = _experiment_result()
    output_path = tmp_path / "experiment_summary.xlsx"
    export_summary_workbook(result, output_path)

    workbook = load_workbook(output_path, read_only=True)
    variant_summary = workbook["variant_summary"]
    fold_metrics = workbook["fold_metrics"]
    holdout_metrics = workbook["holdout_metrics"]
    variant_columns = [cell.value for cell in next(variant_summary.iter_rows(max_row=1))]
    fold_columns = [cell.value for cell in next(fold_metrics.iter_rows(max_row=1))]
    holdout_columns = [cell.value for cell in next(holdout_metrics.iter_rows(max_row=1))]

    assert {
        "variant_order",
        "variant_id",
        "status",
        "entry_node",
        "exit_node",
        "risk_node",
        "entry_params",
        "exit_params",
        "risk_params",
    }.issubset(variant_columns)
    assert {
        "variant_order",
        "fold_order",
        "variant_id",
        "fold_label",
        "variant_status",
        "fold_status",
        "runtime_seconds",
    }.issubset(fold_columns)
    assert {
        "variant_order",
        "variant_id",
        "holdout_label",
        "variant_status",
        "holdout_status",
        "runtime_seconds",
    }.issubset(holdout_columns)


def test_summary_frames_serialize_mapping_metrics_with_non_string_keys() -> None:
    result = _experiment_result()

    frames = build_summary_frames(result)
    variant_summary = frames["variant_summary"]
    fold_metrics = frames["fold_metrics"]

    assert variant_summary["cv_holding_period_distribution"].tolist() == [
        '{"1":2.0}',
        '{"1":2.0}',
    ]
    assert fold_metrics["metric_holding_period_distribution"].tolist() == [
        '{"1":1}',
        '{"1":1}',
        '{"1":1}',
        '{"1":1}',
    ]


def test_summary_frames_serialize_non_finite_prune_details() -> None:
    result = _experiment_result()
    result = replace(
        result,
        prune_records=(
            PruneRecord(
                variant_id=result.experiment.variants[0].variant_id,
                stage="after_cv",
                reason="example",
                metrics_snapshot={"payoff_ratio": float("inf"), "holding_period_distribution": {1: 2}},
            ),
        ),
    )

    frames = build_summary_frames(result)
    details = frames["failures_prunes"]["details"].tolist()

    assert details == ['{"holding_period_distribution":{"1":2},"payoff_ratio":"inf"}']


def test_reporting_preserves_deterministic_variant_and_fold_ordering() -> None:
    result = _experiment_result()

    frames = build_summary_frames(result)
    variant_summary = frames["variant_summary"]
    fold_metrics = frames["fold_metrics"]
    holdout_metrics = frames["holdout_metrics"]

    expected_variant_ids = [variant.variant_id for variant in result.experiment.variants]
    assert variant_summary["variant_id"].tolist() == expected_variant_ids
    assert holdout_metrics["variant_id"].tolist() == expected_variant_ids
    assert fold_metrics["variant_id"].tolist() == [
        expected_variant_ids[0],
        expected_variant_ids[0],
        expected_variant_ids[1],
        expected_variant_ids[1],
    ]
    assert fold_metrics["fold_label"].tolist() == ["fold_a", "fold_b", "fold_a", "fold_b"]


def test_holdout_sheet_keeps_one_row_per_variant_with_pruned_variants_visible() -> None:
    result = _experiment_result()

    holdout_metrics = build_summary_frames(result)["holdout_metrics"]

    assert len(holdout_metrics) == 2
    assert holdout_metrics["holdout_status"].tolist() == ["completed", "not_run"]
    assert holdout_metrics["variant_status"].tolist() == ["completed", "pruned"]


def test_deep_dive_selection_accepts_variant_ids_and_specific_targets() -> None:
    result = _experiment_result()
    variant_id = result.experiment.variants[0].variant_id

    targets = select_deep_dive_targets(
        result,
        selected_variant_ids=(variant_id,),
        selected_folds=("fold_b",),
        include_holdout=True,
    )

    assert [target.variant_id for target in targets] == [variant_id, variant_id]
    assert [(target.phase, target.label) for target in targets] == [
        ("cv", "fold_b"),
        ("holdout", "holdout"),
    ]


def test_deep_dive_defaults_to_single_fold_scope_for_price_plot_inputs() -> None:
    result = _experiment_result()
    data = _data()
    variant_id = result.experiment.variants[0].variant_id

    targets = select_deep_dive_targets(result, selected_variant_ids=(variant_id,))

    assert len(targets) == 1
    target = targets[0]
    assert target.phase == "cv"
    assert target.label == "fold_a"

    price_frame = build_price_plot_frame(result, data, target)
    assert not price_frame.empty
    assert price_frame["ts"].min() >= target.window_start
    assert price_frame["ts"].max() < target.window_end
    assert price_frame["ts"].tolist() == [pd.Timestamp("2024-01-02 00:00:00")]


def test_trade_log_export_creation(tmp_path: Path) -> None:
    result = _experiment_result()
    data = _data()
    variant_id = result.experiment.variants[0].variant_id

    artifacts = export_deep_dive_artifacts(
        result,
        data,
        tmp_path / "deep_dive",
        selected_variant_ids=(variant_id,),
        selected_folds=("fold_a",),
        include_holdout=False,
    )

    assert len(artifacts) == 1
    artifact = artifacts[0]
    assert artifact.target.variant_id == variant_id
    assert artifact.target.label == "fold_a"
    assert artifact.equity_plot_path is not None and artifact.equity_plot_path.exists()
    assert artifact.price_plot_path is not None and artifact.price_plot_path.exists()
    assert artifact.trade_log_path is not None and artifact.trade_log_path.exists()
    assert artifact.target_dir.parent.name == variant_id
    assert artifact.target_dir.name == "cv_fold_a"

    trade_log = pd.read_csv(artifact.trade_log_path)
    assert {
        "variant_id",
        "target_phase",
        "target_label",
        "trade_id",
        "entry_ts",
        "exit_ts",
        "side",
        "qty",
        "entry_price",
        "exit_price",
        "gross_pnl",
        "net_pnl",
        "equity_after_trade",
    }.issubset(trade_log.columns)
    assert trade_log["target_label"].tolist() == ["fold_a"]

    selected_target = select_deep_dive_targets(
        result,
        selected_variant_ids=(variant_id,),
        selected_folds=("fold_a",),
    )[0]
    trade_frame = build_trade_log_frame(result, selected_target)
    assert trade_frame["equity_after_trade"].iloc[0] == 100_000.0
